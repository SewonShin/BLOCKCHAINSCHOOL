# import requests
# from bs4 import BeautifulSoup as bs

# URL = "https://www.fastcampus.co.kr/dev_online_introp/"

# rq = requests.get(URL)
# soup = bs(rq.content, 'html.parser')
# print(soup)





import requests
from bs4 import BeautifulSoup as bs
from konlpy.tag import Okt
from openpyxl import load_workbook as load

URL = "https://www.fastcampus.co.kr/dev_online_introp/"
SAVE_DIR = DIR = "/Users/kamil/Desktop/PROGRAMMING/PYTHON/FastCampusPythonBasic/code14/test.xlsx"

def get_tags():
    tags_list = []
    # div.vc_toggle_content > ul > li
    rq = requests.get(URL)
    print(rq.status_code)
    if rq.status_code == 200:
        soup = bs(rq.content, 'html.parser')

        # (class, id, ...), attr = {key : value, ...}
        li_list = soup.find_all('div', class_ = 'vc_toggle_content')
        # li_list -> 모든 강의 내용, 강의 과제 
        for li in li_list:
            ul = li.select('ul > li')
            for l in ul:
                tags_list.append(l.get_text())
                # print(l.get_text())
        return tags_list
    else:
        raise Exception('응답 코드가 200번이 아닙니다.')

def save_excel(rank):
    wb = load(SAVE_DIR)
    ws = wb.create_sheet('rank')
    idx = 1
    try:
        for k, v in rank[:15]:
            ws['A' + str(idx)] = "{}({})".format(k, v)
            idx += 1
        wb.save(SAVE_DIR)
    except Exception as e:
        print(e)
    finally:
        wb.close() 

def get_rank():
    rank = {}
    tags_list = get_tags()
    ok = Okt()
    for tag in tags_list:
        noun = ok.nouns(tag)
        for n in noun:
            if not (n in rank):
                rank[n] = 0
            rank[n] += 1
    rank = sorted(rank.items(), key =lambda x:x[1], reverse = True)
#    for k, v in rank:
#        print("{}({})".format(k, v), end = ' ')

if __name__ == "__main__":
    get_rank()










