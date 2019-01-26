from openpyxl import load_workbook as load

DIR = '/Users/kamil/Desktop/PROGRAMMING/PYTHON/FastCampusPythonBasic/code09/test.xlsx'

wb = load(DIR)

# ws = wb.create_sheet('test')
ws = wb['test']
# ws['A1'] = "제목1"
# ws['B1'] = "제목2"
a1 = ws['A2'].value
a2 = ws['B2'].value
print(a1, a2)
# wb.save(DIR)
wb.close()
